"""Gerador de relatório Excel (.xlsx) com formatação farol para compliance.

Converte uma lista de ComplianceReport em bytes Excel formatados com cores
verde/amarelo/vermelho baseadas no status de compliance e na confidence
de cada regra.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from brand_watchdog.models.dataclasses import (
    ComplianceReport,
    ComplianceRuleResult,
)


class ExcelComplianceReportGenerator:
    """Gera relatório Excel (.xlsx) com formatação farol a partir de
    dados de compliance.
    """

    # Colunas fixas na ordem definida
    COLUMNS: list[str] = [
        "URL",
        "facilitator_role",
        "logo_application",
        "logo_effects",
        "content_separation",
        "naming_pricing",
        "kv_integrity",
        "Status",
    ]

    # Regras na ordem das colunas B-G
    RULE_IDS: list[str] = [
        "facilitator_role",
        "logo_application",
        "logo_effects",
        "content_separation",
        "naming_pricing",
        "kv_integrity",
    ]

    # Cores de farol (sem prefixo alpha — openpyxl usa RRGGBB de 6 dígitos)
    COLOR_GREEN: str = "00B050"
    COLOR_YELLOW: str = "FFFF00"
    COLOR_RED: str = "FF0000"

    # Thresholds de confidence
    THRESHOLD_HIGH: int = 80
    THRESHOLD_MID: int = 60

    # Limite de regras NOT_APPLICABLE para classificar "Not detected"
    NOT_APPLICABLE_THRESHOLD: int = 4

    def generate(self, reports: list[ComplianceReport]) -> bytes:
        """Gera arquivo Excel em memória.

        Args:
            reports: Lista de relatórios de compliance.

        Returns:
            Bytes do arquivo .xlsx pronto para anexar.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Compliance Report"

        # Header row
        ws.append(self.COLUMNS)

        # Data rows
        for report in reports:
            status_text = self._determine_status_text(report)
            is_not_detected = status_text == "Not detected"

            # Mapear rule_results por rule_id para acesso O(1)
            rule_map: dict[str, ComplianceRuleResult] = {
                rr.rule_id: rr for rr in report.rule_results
            }

            # Montar linha de dados
            row_values: list[str] = [report.target_url]

            for rule_id in self.RULE_IDS:
                rule_result = rule_map.get(rule_id)
                row_values.append(self._get_rule_cell_value(rule_result))

            row_values.append(status_text)
            ws.append(row_values)

            # Aplicar fills de cor na linha recém-inserida
            current_row = ws.max_row

            # Fills nas colunas de regras (B-G, índices 2-7)
            for col_idx, rule_id in enumerate(self.RULE_IDS, start=2):
                rule_result = rule_map.get(rule_id)
                fill = self._get_rule_fill(rule_result, is_not_detected)
                ws.cell(row=current_row, column=col_idx).fill = fill

            # Fill na coluna Status (H, índice 8)
            status_fill = self._get_status_fill(status_text)
            ws.cell(row=current_row, column=8).fill = status_fill

        # Salvar em memória
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _determine_status_text(self, report: ComplianceReport) -> str:
        """Determina o texto da coluna Status.

        Lógica:
        - non_compliant → "NON COMPLIANT"
        - compliant + ≥4 regras NOT_APPLICABLE → "Not detected"
        - compliant + <4 NOT_APPLICABLE → "COMPLIANT"

        Returns:
            "NON COMPLIANT", "COMPLIANT", ou "Not detected".
        """
        if report.overall_status == "non_compliant":
            return "NON COMPLIANT"

        # Contar regras com status NOT_APPLICABLE
        na_count = sum(
            1
            for rr in report.rule_results
            if rr.status == "NOT_APPLICABLE"
        )

        if na_count >= self.NOT_APPLICABLE_THRESHOLD:
            return "Not detected"

        return "COMPLIANT"

    def _get_rule_cell_value(
        self, rule_result: ComplianceRuleResult | None
    ) -> str:
        """Retorna valor da célula de regra.

        Args:
            rule_result: Resultado da regra, ou None se não encontrado.

        Returns:
            "N/A" para NOT_APPLICABLE ou regra ausente, ou "{confidence}%".
        """
        if rule_result is None:
            return "N/A"

        if rule_result.status == "NOT_APPLICABLE":
            return "N/A"

        return f"{rule_result.confidence}%"

    def _get_status_fill(self, status_text: str) -> PatternFill:
        """Retorna fill color para a célula de Status.

        - "COMPLIANT" → verde
        - "Not detected" → verde
        - "NON COMPLIANT" → vermelho
        """
        if status_text in ("COMPLIANT", "Not detected"):
            return PatternFill(
                start_color=self.COLOR_GREEN,
                end_color=self.COLOR_GREEN,
                fill_type="solid",
            )

        # "NON COMPLIANT"
        return PatternFill(
            start_color=self.COLOR_RED,
            end_color=self.COLOR_RED,
            fill_type="solid",
        )

    def _get_rule_fill(
        self,
        rule_result: ComplianceRuleResult | None,
        is_not_detected: bool,
    ) -> PatternFill:
        """Retorna fill color para célula de regra.

        Se o site é "Not detected", todas as células de regra ficam verdes
        (override). Caso contrário, aplica thresholds de confidence:
        - confidence >= 80 → verde
        - 60 <= confidence < 80 → amarelo
        - confidence < 60 → vermelho

        Args:
            rule_result: Resultado da regra, ou None se não encontrado.
            is_not_detected: True se o status do site é "Not detected".

        Returns:
            PatternFill com a cor apropriada.
        """
        # Override verde para sites "Not detected"
        if is_not_detected:
            return PatternFill(
                start_color=self.COLOR_GREEN,
                end_color=self.COLOR_GREEN,
                fill_type="solid",
            )

        # Se regra ausente, confidence é tratada como 0 (< 60 → vermelho)
        confidence = rule_result.confidence if rule_result else 0

        if confidence >= self.THRESHOLD_HIGH:
            color = self.COLOR_GREEN
        elif confidence >= self.THRESHOLD_MID:
            color = self.COLOR_YELLOW
        else:
            color = self.COLOR_RED

        return PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid",
        )
