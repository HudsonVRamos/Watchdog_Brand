"""Property tests para coloração de células de regras baseada em confidence.

# Feature: excel-compliance-report, Property 3: Rule cell confidence-based coloring

**Validates: Requirements 3.1, 3.2, 3.3, 3.5**

Property 3: Para qualquer célula de regra em um site cujo Overall_Status NÃO é
"Not detected", a cor de fundo deve ser: verde (#00B050) se confidence >= 80,
amarelo (#FFFF00) se 60 <= confidence < 80, e vermelho (#FF0000) se
confidence < 60. Isso inclui regras NOT_APPLICABLE cujo confidence segue os
mesmos thresholds.
"""

from __future__ import annotations

from io import BytesIO

from hypothesis import given, settings, HealthCheck, assume
from hypothesis import strategies as st
from openpyxl import load_workbook

from brand_watchdog.reports.excel_compliance_report_generator import (
    ExcelComplianceReportGenerator,
)

from tests.property.strategies import compliance_report


# Cores esperadas (ARGB 8 chars como openpyxl retorna)
_GREEN = "0000B050"
_YELLOW = "00FFFF00"
_RED = "00FF0000"

# Ordem das regras mapeadas às colunas B-G
_RULE_IDS = [
    "facilitator_role",
    "logo_application",
    "logo_effects",
    "content_separation",
    "naming_pricing",
    "kv_integrity",
]

# Threshold para "Not detected"
_NOT_APPLICABLE_THRESHOLD = 4


def _is_not_detected(report) -> bool:
    """Verifica se um report é classificado como 'Not detected'."""
    if report.overall_status != "compliant":
        return False
    na_count = sum(
        1 for rr in report.rule_results if rr.status == "NOT_APPLICABLE"
    )
    return na_count >= _NOT_APPLICABLE_THRESHOLD


def _expected_color(confidence: int) -> str:
    """Retorna a cor ARGB esperada baseada no confidence."""
    if confidence >= 80:
        return _GREEN
    elif confidence >= 60:
        return _YELLOW
    else:
        return _RED


class TestRuleCellConfidenceBasedColoring:
    """Property 3: Rule cell confidence-based coloring.

    Para sites NÃO classificados como "Not detected", a cor de cada célula
    de regra deve refletir os thresholds de confidence:
    - confidence >= 80 → verde (#00B050)
    - 60 <= confidence < 80 → amarelo (#FFFF00)
    - confidence < 60 → vermelho (#FF0000)

    **Validates: Requirements 3.1, 3.2, 3.3, 3.5**
    """

    @settings(
        max_examples=100,
        deadline=None,
        suppress_health_check=[HealthCheck.too_slow],
    )
    @given(reports=st.lists(compliance_report(), min_size=1, max_size=10))
    def test_rule_cells_colored_by_confidence_thresholds(
        self,
        reports,
    ):
        """Para cada site não 'Not detected', cada célula de regra deve ter
        fill baseado nos thresholds de confidence."""
        # Precisamos de pelo menos um report que NÃO seja "Not detected"
        assume(any(not _is_not_detected(r) for r in reports))

        generator = ExcelComplianceReportGenerator()
        excel_bytes = generator.generate(reports)

        wb = load_workbook(filename=BytesIO(excel_bytes))
        ws = wb.active

        for row_idx, report in enumerate(reports, start=2):
            # Pular sites "Not detected" (domínio da Property 4)
            if _is_not_detected(report):
                continue

            # Mapear rule_results por rule_id
            rule_map = {rr.rule_id: rr for rr in report.rule_results}

            # Verificar cada coluna de regra (B-G, índices 2-7)
            for col_idx, rule_id in enumerate(_RULE_IDS, start=2):
                cell = ws.cell(row=row_idx, column=col_idx)
                rule_result = rule_map.get(rule_id)

                # Determinar confidence esperado
                if rule_result is None:
                    confidence = 0
                else:
                    confidence = rule_result.confidence

                expected = _expected_color(confidence)
                actual = cell.fill.fgColor.rgb

                assert actual == expected, (
                    f"Row {row_idx}, Col {col_idx} ({rule_id}): "
                    f"confidence={confidence}, "
                    f"expected fill={expected}, got={actual}. "
                    f"Rule status={rule_result.status if rule_result else 'None'}"
                )
