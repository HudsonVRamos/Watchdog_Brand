"""Property test para round-trip de integridade de dados do Excel.

# Feature: excel-compliance-report, Property 1: Round-trip data integrity

**Validates: Requirements 6.1, 6.2, 1.1, 1.3, 1.4, 1.5, 1.6, 1.7**

Para qualquer lista não-vazia de ComplianceReport, gerar bytes Excel e
carregá-los de volta com openpyxl deve produzir: mesmo número de linhas,
URLs preservadas na ordem, Status correto, e valores de regras corretos.
"""

from __future__ import annotations

from io import BytesIO

from hypothesis import HealthCheck, given, settings
from hypothesis import strategies as st
from openpyxl import load_workbook

from brand_watchdog.reports.excel_compliance_report_generator import (
    ExcelComplianceReportGenerator,
)

from .strategies import compliance_report


_PBT_SETTINGS = settings(
    max_examples=100,
    suppress_health_check=[HealthCheck.too_slow],
    deadline=None,
)

# Constantes alinhadas com o gerador
_RULE_IDS = ExcelComplianceReportGenerator.RULE_IDS
_NOT_APPLICABLE_THRESHOLD = (
    ExcelComplianceReportGenerator.NOT_APPLICABLE_THRESHOLD
)


def _expected_status_text(report) -> str:
    """Calcula o texto esperado da coluna Status para um report."""
    if report.overall_status == "non_compliant":
        return "NON COMPLIANT"

    na_count = sum(
        1
        for rr in report.rule_results
        if rr.status == "NOT_APPLICABLE"
    )

    if na_count >= _NOT_APPLICABLE_THRESHOLD:
        return "Not detected"

    return "COMPLIANT"


def _expected_rule_cell_value(rule_result) -> str:
    """Calcula o valor esperado da célula de regra."""
    if rule_result is None:
        return "N/A"
    if rule_result.status == "NOT_APPLICABLE":
        return "N/A"
    return f"{rule_result.confidence}%"


class TestExcelReportRoundTripDataIntegrity:
    """Property 1: Round-trip data integrity.

    **Validates: Requirements 6.1, 6.2, 1.1, 1.3, 1.4, 1.5, 1.6, 1.7**
    """

    @_PBT_SETTINGS
    @given(
        reports=st.lists(compliance_report(), min_size=1, max_size=10)
    )
    def test_roundtrip_preserves_row_count(self, reports) -> None:
        """Número de data rows no Excel == len(reports)."""
        generator = ExcelComplianceReportGenerator()
        excel_bytes = generator.generate(reports)

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active

        # Row 1 é o header, data rows começam em row 2
        data_rows = ws.max_row - 1
        assert data_rows == len(reports), (
            f"Esperado {len(reports)} data rows, encontrado {data_rows}"
        )

    @_PBT_SETTINGS
    @given(
        reports=st.lists(compliance_report(), min_size=1, max_size=10)
    )
    def test_roundtrip_preserves_urls_in_order(self, reports) -> None:
        """Column A (URL) preserva target_url na ordem de inserção."""
        generator = ExcelComplianceReportGenerator()
        excel_bytes = generator.generate(reports)

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active

        for i, report in enumerate(reports):
            row_idx = i + 2  # header na row 1
            cell_value = ws.cell(row=row_idx, column=1).value
            assert cell_value == report.target_url, (
                f"Row {row_idx}: URL esperada {report.target_url!r}, "
                f"encontrada {cell_value!r}"
            )

    @_PBT_SETTINGS
    @given(
        reports=st.lists(compliance_report(), min_size=1, max_size=10)
    )
    def test_roundtrip_preserves_status_text(self, reports) -> None:
        """Column H (Status) exibe texto correto derivado do report."""
        generator = ExcelComplianceReportGenerator()
        excel_bytes = generator.generate(reports)

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active

        for i, report in enumerate(reports):
            row_idx = i + 2
            cell_value = ws.cell(row=row_idx, column=8).value
            expected = _expected_status_text(report)
            assert cell_value == expected, (
                f"Row {row_idx}: Status esperado {expected!r}, "
                f"encontrado {cell_value!r} "
                f"(overall_status={report.overall_status!r})"
            )

    @_PBT_SETTINGS
    @given(
        reports=st.lists(compliance_report(), min_size=1, max_size=10)
    )
    def test_roundtrip_preserves_rule_values(self, reports) -> None:
        """Colunas B-G preservam valores de regras corretos."""
        generator = ExcelComplianceReportGenerator()
        excel_bytes = generator.generate(reports)

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active

        for i, report in enumerate(reports):
            row_idx = i + 2

            # Mapear rule_results por rule_id
            rule_map = {rr.rule_id: rr for rr in report.rule_results}

            for col_offset, rule_id in enumerate(_RULE_IDS):
                col_idx = col_offset + 2  # Coluna B = 2
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                rule_result = rule_map.get(rule_id)
                expected = _expected_rule_cell_value(rule_result)
                assert cell_value == expected, (
                    f"Row {row_idx}, Col {col_idx} ({rule_id}): "
                    f"esperado {expected!r}, encontrado {cell_value!r}"
                )
