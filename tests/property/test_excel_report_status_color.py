"""Property test: Status cell color correctness.

Feature: excel-compliance-report, Property 2: Status cell color correctness

Verifica que a cor de fundo da célula Status (coluna H) é verde (#00B050)
para "COMPLIANT" e "Not detected", e vermelha (#FF0000) para "NON COMPLIANT".

**Validates: Requirements 2.1, 2.2, 2.3**
"""

from __future__ import annotations

from io import BytesIO

from hypothesis import HealthCheck, given, settings
from hypothesis import strategies as st
from openpyxl import load_workbook

from brand_watchdog.reports.excel_compliance_report_generator import (
    ExcelComplianceReportGenerator,
)
from tests.property.strategies import compliance_report


# Cores esperadas em formato ARGB de 8 caracteres (alpha "00" + RGB)
EXPECTED_GREEN_ARGB = "0000B050"
EXPECTED_RED_ARGB = "00FF0000"


@given(reports=st.lists(compliance_report(), min_size=1, max_size=10))
@settings(max_examples=100, suppress_health_check=[HealthCheck.too_slow])
def test_status_cell_color_correctness(
    reports: list,
) -> None:
    """Para qualquer ComplianceReport, a cor da célula Status deve ser:
    - Verde (#00B050) quando o texto é "COMPLIANT" ou "Not detected"
    - Vermelho (#FF0000) quando o texto é "NON COMPLIANT"
    """
    generator = ExcelComplianceReportGenerator()
    excel_bytes = generator.generate(reports)

    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active

    # Coluna H = índice 8 (Status)
    status_col_index = 8

    # Iterar pelas linhas de dados (começando na linha 2, pois linha 1 é header)
    for row_idx in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row_idx, column=status_col_index)
        status_text = status_cell.value

        # Verificar fill_type é "solid"
        assert status_cell.fill.fill_type == "solid", (
            f"Linha {row_idx}: fill_type da célula Status deveria ser "
            f"'solid', mas é '{status_cell.fill.fill_type}'"
        )

        # Obter cor RGB (formato ARGB de 8 caracteres no openpyxl)
        actual_color = status_cell.fill.fgColor.rgb

        if status_text in ("COMPLIANT", "Not detected"):
            assert actual_color == EXPECTED_GREEN_ARGB, (
                f"Linha {row_idx}: Status '{status_text}' deveria ter "
                f"cor verde ({EXPECTED_GREEN_ARGB}), mas tem '{actual_color}'"
            )
        elif status_text == "NON COMPLIANT":
            assert actual_color == EXPECTED_RED_ARGB, (
                f"Linha {row_idx}: Status '{status_text}' deveria ter "
                f"cor vermelha ({EXPECTED_RED_ARGB}), mas tem '{actual_color}'"
            )
        else:
            # Status inesperado — falha do teste
            raise AssertionError(
                f"Linha {row_idx}: Status inesperado: '{status_text}'. "
                f"Valores válidos: 'COMPLIANT', 'NON COMPLIANT', 'Not detected'"
            )
