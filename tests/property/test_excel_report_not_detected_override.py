"""Property test para override verde em sites "Not detected".

# Feature: excel-compliance-report, Property 4: "Not detected" green override

**Validates: Requirements 3.4**

Para qualquer site classificado como "Not detected" (compliant + ≥4 regras
NOT_APPLICABLE), TODAS as células de regras (colunas B-G) devem ter fill
verde (#00B050), independente dos valores de confidence individuais.
"""

from __future__ import annotations

from io import BytesIO

from hypothesis import HealthCheck, given, settings
from hypothesis import strategies as st
from openpyxl import load_workbook

from brand_watchdog.models.dataclasses import (
    COMPLIANCE_RULES,
    ComplianceReport,
    ComplianceRuleResult,
)
from brand_watchdog.reports.excel_compliance_report_generator import (
    ExcelComplianceReportGenerator,
)

from .strategies import analyzed_at_datetime


_PBT_SETTINGS = settings(
    max_examples=100,
    deadline=None,
    suppress_health_check=[
        HealthCheck.too_slow,
        HealthCheck.filter_too_much,
    ],
)

# Cor verde esperada (ARGB 8 caracteres com prefixo alpha "00")
_GREEN_ARGB = "0000B050"

# Constantes do gerador
_RULE_IDS = ExcelComplianceReportGenerator.RULE_IDS
_NOT_APPLICABLE_THRESHOLD = (
    ExcelComplianceReportGenerator.NOT_APPLICABLE_THRESHOLD
)


@st.composite
def not_detected_report(draw: st.DrawFn) -> ComplianceReport:
    """Gera um ComplianceReport que se classifica como "Not detected".

    Requisitos para "Not detected":
    - overall_status == "compliant" (nenhuma regra FAIL)
    - >= 4 regras com status NOT_APPLICABLE

    Estratégia: gera entre 4 e 6 regras como NOT_APPLICABLE,
    as restantes como PASS (nunca FAIL para manter compliant).
    Confidence é aleatória para todas as regras.
    """
    na_count = draw(st.integers(min_value=4, max_value=6))

    # Selecionar quais regras serão NOT_APPLICABLE
    all_indices = list(range(6))
    na_indices_set = set(
        draw(
            st.lists(
                st.sampled_from(all_indices),
                min_size=na_count,
                max_size=na_count,
                unique=True,
            )
        )
    )

    rule_results = []
    for i, rule_id in enumerate(COMPLIANCE_RULES):
        if i in na_indices_set:
            status = "NOT_APPLICABLE"
        else:
            status = "PASS"

        confidence = draw(st.integers(min_value=0, max_value=100))
        description = draw(
            st.text(
                alphabet=st.characters(
                    whitelist_categories=("L", "N", "P", "Z"),
                ),
                min_size=1,
                max_size=50,
            )
        )
        rule_results.append(
            ComplianceRuleResult(
                rule_id=rule_id,
                status=status,
                confidence=confidence,
                description=description,
            )
        )

    target_url = draw(
        st.from_regex(
            r"https://[a-z]{3,10}\.[a-z]{2,5}/[a-z]{1,10}",
            fullmatch=True,
        )
    )
    analyzed_at = draw(analyzed_at_datetime())

    return ComplianceReport(
        target_url=target_url,
        analyzed_at=analyzed_at,
        overall_status="compliant",
        rule_results=rule_results,
    )


class TestExcelReportNotDetectedGreenOverride:
    """Property 4: "Not detected" green override.

    **Validates: Requirements 3.4**

    Para qualquer site classificado como "Not detected", TODAS as
    células de regras devem ter fill verde, independente de confidence.
    """

    @_PBT_SETTINGS
    @given(
        reports=st.lists(
            not_detected_report(), min_size=1, max_size=5
        )
    )
    def test_not_detected_all_rule_cells_are_green(
        self, reports
    ) -> None:
        """Todas as células de regra (B-G) de sites "Not detected"
        devem ter fill verde (#00B050)."""
        generator = ExcelComplianceReportGenerator()
        excel_bytes = generator.generate(reports)

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active

        for i, report in enumerate(reports):
            row_idx = i + 2  # header na row 1

            # Confirmar que realmente é "Not detected"
            status_cell = ws.cell(row=row_idx, column=8).value
            assert status_cell == "Not detected", (
                f"Row {row_idx}: Esperado 'Not detected', "
                f"encontrado {status_cell!r}"
            )

            # Verificar fill verde em TODAS as colunas de regras (B-G)
            for col_idx in range(2, 8):  # colunas 2 a 7
                cell = ws.cell(row=row_idx, column=col_idx)
                fill = cell.fill

                assert fill.fill_type == "solid", (
                    f"Row {row_idx}, Col {col_idx}: "
                    f"fill_type esperado 'solid', "
                    f"encontrado {fill.fill_type!r}"
                )

                actual_color = str(fill.fgColor.rgb)
                assert actual_color == _GREEN_ARGB, (
                    f"Row {row_idx}, Col {col_idx} "
                    f"({_RULE_IDS[col_idx - 2]}): "
                    f"cor esperada {_GREEN_ARGB}, "
                    f"encontrada {actual_color}. "
                    f"confidence={report.rule_results[col_idx - 2].confidence}"
                )

    @_PBT_SETTINGS
    @given(
        reports=st.lists(
            not_detected_report(), min_size=1, max_size=5
        )
    )
    def test_not_detected_override_ignores_confidence(
        self, reports
    ) -> None:
        """O override verde deve aplicar-se independente do valor de
        confidence (0-100), incluindo valores baixos que normalmente
        seriam amarelo ou vermelho."""
        generator = ExcelComplianceReportGenerator()
        excel_bytes = generator.generate(reports)

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active

        for i, report in enumerate(reports):
            row_idx = i + 2
            rule_map = {
                rr.rule_id: rr for rr in report.rule_results
            }

            for col_offset, rule_id in enumerate(_RULE_IDS):
                col_idx = col_offset + 2
                cell = ws.cell(row=row_idx, column=col_idx)
                rule_result = rule_map.get(rule_id)

                # Mesmo regras com confidence < 60 (normalmente
                # vermelho) devem ser verdes por causa do override
                actual_color = str(cell.fill.fgColor.rgb)
                assert actual_color == _GREEN_ARGB, (
                    f"Row {row_idx}, regra '{rule_id}': "
                    f"Override verde falhou. "
                    f"confidence={rule_result.confidence if rule_result else 'N/A'}, "
                    f"status={rule_result.status if rule_result else 'missing'}, "
                    f"cor encontrada={actual_color}"
                )
