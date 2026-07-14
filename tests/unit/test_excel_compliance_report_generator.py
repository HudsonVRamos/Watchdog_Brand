"""Testes unitários para ExcelComplianceReportGenerator.

Validates: Requirements 1.1, 1.2, 1.6, 1.7, 1.8, 6.1, 6.3
"""

from __future__ import annotations

import pytest
from io import BytesIO
from datetime import datetime, timezone
from typing import Optional, List

from openpyxl import load_workbook

from brand_watchdog.reports.excel_compliance_report_generator import (
    ExcelComplianceReportGenerator,
)
from brand_watchdog.models.dataclasses import (
    ComplianceReport,
    ComplianceRuleResult,
)


# --- Helpers / Fixtures ---


def _make_rule(rule_id: str, status: str = "PASS", confidence: int = 85):
    """Cria um ComplianceRuleResult para testes."""
    return ComplianceRuleResult(
        rule_id=rule_id,
        status=status,
        confidence=confidence,
        description="Test description",
    )


def _make_report(
    url: str = "https://example.com",
    overall_status: str = "compliant",
    rules: Optional[List[ComplianceRuleResult]] = None,
) -> ComplianceReport:
    """Cria um ComplianceReport para testes."""
    if rules is None:
        rules = []
    return ComplianceReport(
        target_url=url,
        analyzed_at=datetime(2025, 1, 15, 10, 0, 0, tzinfo=timezone.utc),
        overall_status=overall_status,
        rule_results=rules,
    )


@pytest.fixture
def generator():
    """Fixture do gerador Excel."""
    return ExcelComplianceReportGenerator()


# --- Tests ---


class TestEmptyReports:
    """Testa geração com lista vazia de reports."""

    def test_empty_reports_generates_valid_excel_with_header_only(
        self, generator: ExcelComplianceReportGenerator
    ):
        """Lista vazia gera xlsx válido com apenas header row.

        Validates: Requirements 1.8, 6.3
        """
        result = generator.generate([])

        # Deve ser bytes válidos
        assert isinstance(result, bytes)
        assert len(result) > 0

        # Deve ser carregável por openpyxl
        wb = load_workbook(BytesIO(result))
        ws = wb.active

        # Apenas 1 linha (header), sem dados
        assert ws.max_row == 1

        # Header deve estar preenchido
        header_values = [ws.cell(row=1, column=c).value for c in range(1, 9)]
        assert header_values == ExcelComplianceReportGenerator.COLUMNS


class TestHeaderColumnOrder:
    """Testa ordem fixa das colunas no header."""

    def test_header_column_order_is_fixed(
        self, generator: ExcelComplianceReportGenerator
    ):
        """Header row tem colunas na ordem fixa especificada.

        Validates: Requirements 1.2
        """
        report = _make_report(
            rules=[_make_rule("facilitator_role", "PASS", 90)]
        )
        result = generator.generate([report])

        wb = load_workbook(BytesIO(result))
        ws = wb.active

        expected_columns = [
            "URL",
            "facilitator_role",
            "logo_application",
            "logo_effects",
            "content_separation",
            "naming_pricing",
            "kv_integrity",
            "Status",
        ]

        header_values = [ws.cell(row=1, column=c).value for c in range(1, 9)]
        assert header_values == expected_columns


class TestMissingRuleDisplaysNA:
    """Testa que regra ausente em rule_results exibe N/A."""

    def test_missing_rule_displays_na(
        self, generator: ExcelComplianceReportGenerator
    ):
        """Regras ausentes no rule_results devem exibir 'N/A' na célula.

        Validates: Requirements 1.7
        """
        # Report com apenas 3 regras (faltam logo_effects, content_separation, kv_integrity)
        report = _make_report(
            rules=[
                _make_rule("facilitator_role", "PASS", 90),
                _make_rule("logo_application", "PASS", 75),
                _make_rule("naming_pricing", "FAIL", 40),
            ]
        )
        result = generator.generate([report])

        wb = load_workbook(BytesIO(result))
        ws = wb.active

        # Linha de dados é row 2
        # Colunas das regras ausentes:
        # logo_effects = coluna D (4)
        # content_separation = coluna E (5)
        # kv_integrity = coluna G (7)
        assert ws.cell(row=2, column=4).value == "N/A"  # logo_effects
        assert ws.cell(row=2, column=5).value == "N/A"  # content_separation
        assert ws.cell(row=2, column=7).value == "N/A"  # kv_integrity

        # Regras presentes NÃO devem ser N/A
        assert ws.cell(row=2, column=2).value == "90%"  # facilitator_role
        assert ws.cell(row=2, column=3).value == "75%"  # logo_application
        assert ws.cell(row=2, column=6).value == "40%"  # naming_pricing


class TestConfidenceFormatting:
    """Testa formatação de confidence como inteiro percentual."""

    def test_confidence_formatted_as_integer_percentage(
        self, generator: ExcelComplianceReportGenerator
    ):
        """Confidence deve ser formatada como '{confidence}%' (ex: '85%').

        Validates: Requirements 1.6
        """
        report = _make_report(
            rules=[_make_rule("facilitator_role", "PASS", 85)]
        )
        result = generator.generate([report])

        wb = load_workbook(BytesIO(result))
        ws = wb.active

        # facilitator_role está na coluna B (2)
        cell_value = ws.cell(row=2, column=2).value
        assert cell_value == "85%"


class TestNotApplicableDisplaysNA:
    """Testa que regras NOT_APPLICABLE exibem N/A."""

    def test_not_applicable_rule_displays_na(
        self, generator: ExcelComplianceReportGenerator
    ):
        """Regra com status NOT_APPLICABLE deve exibir 'N/A', não a confidence.

        Validates: Requirements 1.7
        """
        report = _make_report(
            rules=[
                _make_rule("facilitator_role", "NOT_APPLICABLE", 50),
            ]
        )
        result = generator.generate([report])

        wb = load_workbook(BytesIO(result))
        ws = wb.active

        # Deve exibir "N/A" e NÃO "50%"
        cell_value = ws.cell(row=2, column=2).value
        assert cell_value == "N/A"
        assert cell_value != "50%"


class TestOutputIsValidXlsx:
    """Testa que output é bytes válidos carregáveis por openpyxl."""

    def test_output_is_valid_loadable_xlsx_bytes(
        self, generator: ExcelComplianceReportGenerator
    ):
        """Output deve ser bytes válidos que openpyxl consegue carregar.

        Validates: Requirements 6.1, 6.3
        """
        report = _make_report(
            rules=[
                _make_rule("facilitator_role", "PASS", 90),
                _make_rule("logo_application", "FAIL", 55),
                _make_rule("logo_effects", "NOT_APPLICABLE", 0),
                _make_rule("content_separation", "PASS", 70),
                _make_rule("naming_pricing", "PASS", 80),
                _make_rule("kv_integrity", "PASS", 95),
            ]
        )
        result = generator.generate([report])

        # Deve ser bytes
        assert isinstance(result, bytes)

        # Deve ser carregável sem exceção
        wb = load_workbook(BytesIO(result))

        # Deve ter sheet ativa
        assert wb.active is not None

        # Sheet deve ter título esperado
        assert wb.active.title == "Compliance Report"
